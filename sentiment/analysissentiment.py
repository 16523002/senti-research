from django.contrib.staticfiles.storage import staticfiles_storage
from repository import models
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook, load_workbook
import os

url = os.path.join(os.path.dirname(os.path.dirname(__file__)),'sentiment/inset.xlsx')

# url = staticfiles_storage.url('sentiment/inset.xlsx')
inSetLexicon = load_workbook(url)
negatif = inSetLexicon['negatif']
positif = inSetLexicon['positif']
# tokenization
from nltk.tokenize import sent_tokenize, word_tokenize

# stemming
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
factory = StemmerFactory()
stemmer = factory.create_stemmer()

#stopword removal
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()


# =================================================================================================================
# ============================================== TEXT PRE-PROCESSING ==============================================
# =================================================================================================================

# -------------STOPWORD REMOVAL-------------
def stopwordRemoval(data):
    hasil = []                                                                              #hasil = result
    hasil2 = []
    for kata in data:
        hasil.append(stopword.remove(kata))
    for kata in hasil:
        if kata != '':
            hasil2.append(kata) 
    return hasil2

# -------------PUNCTUATION REMOVAL & CASE CONVERSION-------------
def punctuationRemoval(data):
    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    hasil = []
    for kata in data:
        if kata not in punctuations:
            hasil.append(kata.lower())                                                      #kata = word
    return hasil

# -------------STEMMING-------------
def stemmingWord(data):
    hasil = []
    for kata in data:
        hasil.append(stemmer.stem(kata))
    return hasil

# -------------TOKENIZATION-------------
def tokenization(data):
    hasil = word_tokenize(data)
    return hasil

# -------------JOIN STRING N-GRAM ALL-------------
def nGramAll(hasilPraproses):
    hasilNGram1 = nGram1(hasilPraproses)
    hasilNGram2 = nGram2(hasilPraproses)
    hasilNGram3 = nGram3(hasilPraproses)
    return hasilNGram1, hasilNGram2, hasilNGram3

# -------------JOIN STRING N-GRAM = 1-------------
def nGram1(hasilPraproses):                                                                 #hasilPraproses = Pre-proecessing Result
    return hasilPraproses                                                                   #hasilNGram = NGram Result

# --------------------------
def nGram2(hasilPraproses):                                                                 
    tempKata = ''
    hasilNGram2 = []

    total = int(len(hasilPraproses))

    for i in range(0, total):
        if (i == total-1):
            return hasilNGram2                                                              
        else:
            tempKata = hasilPraproses[i] + ' ' + hasilPraproses[i+1]
            hasilNGram2.append(tempKata)
    
    return hasilNGram2

# -------------JOIN STRING N-GRAM = 3-------------
def nGram3(hasilPraproses):
    tempKata = ''
    hasilNGram3 = []

    total = int(len(hasilPraproses))

    if (total >= 3):
        for j in range(0, total):
            if (j == total-2):
                return hasilNGram3
            else:
                tempKata = hasilPraproses[j] + ' ' + hasilPraproses[j+1] + ' ' + hasilPraproses[j+2] 
                hasilNGram3.append(tempKata)
   
    return hasilNGram3

# -------------FIND SENTIWORD BEFORE STEMMING-------------
def sentiWordBeforeStem(hasilToken):
    arrPositifNGram3 = []
    arrNegatifNGram3 = []
    arrPositifNGram2 = []
    arrNegatifNGram2 = []
    arrPositifNGram1Strip = []
    arrNegatifNGram1Strip = []
    hasilNGramPositif = []
    hasilNGramNegatif = []

    hasilNGram1, hasilNGram2, hasilNGram3 = nGramAll(hasilToken)

    # FIND SENTIWORD N-GRAM3 DATA
    for kata in hasilNGram3:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram3.append([sentiWordPositif,weightPositif])
        for i in range(2, 6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram3.append([sentiWordNegatif,weightNegatif])
        
    # REMOVE N-GRAM3 WORD WITHIN NGRAM1 AND N-GRAM2 DATA (AVOID DOUBLE DATA)
    if (len(arrPositifNGram3) > 0):
        for kata in arrPositifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break
                            
    if (len(arrNegatifNGram3) > 0):
        for kata in arrNegatifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break

    # FIND SENTIWORD N-GRAM2 DATA
    for kata in hasilNGram2:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram2.append([sentiWordPositif,weightPositif])
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram2.append([sentiWordNegatif,weightNegatif])
    
    # REMOVE N-GRAM2 WORD WITHIN N-GRAM1 DATA (AVOID DOUBLE DATA)
    if (len(arrPositifNGram2) > 0):
        for kata in arrPositifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    if (len(arrNegatifNGram2) > 0):
        for kata in arrNegatifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    
    # FIND SENTIWORD N-GRAM1 DATA WITH HYPHENS (EX: BERULANG-ULANG)
    for kata in hasilNGram1:
        if ('-' in kata):
            for i in range(2,3611):
                if (kata == positif.cell(row=i, column=1).value) :
                    sentiWordPositif = (positif.cell(row=i, column=1).value)
                    weightPositif = (positif.cell(row=i, column=2).value)
                    arrPositifNGram1Strip.append([sentiWordPositif,weightPositif])
            for i in range(2,6611):
                if (kata == negatif.cell(row=i, column=1).value) :
                    sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                    weightNegatif = (negatif.cell(row=i, column=2).value)
                    arrNegatifNGram1Strip.append([sentiWordNegatif,weightNegatif])

    # REMOVE N-GRAM1 WORD DATA WITH HYPHENS (AVOID DOUBLE DATA)
    if (len(arrPositifNGram1Strip) > 0):
        for kata in arrPositifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break
    if (len(arrNegatifNGram1Strip) > 0):
        for kata in arrNegatifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break

    for kata in arrPositifNGram1Strip:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram1Strip:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram2:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram2:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram3:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram3:
        hasilNGramNegatif.append(kata)
    
    return hasilNGramPositif, hasilNGramNegatif, hasilNGram1


# -------------FIND SENTIWORD AFTER STEMMING-------------
def sentiWordAfterStem(hasilPraproses, paramNGramPositif, paramNGramNegatif):
    arrPositif = []
    arrNegatif = []

    # FIND SENTIWORD WITH N-GRAM = 1 
    for kata in hasilPraproses:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositif.append([sentiWordPositif,weightPositif])
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatif.append([sentiWordNegatif,weightNegatif])

    for kata in paramNGramPositif:
        arrPositif.append(kata)
    for kata in paramNGramNegatif:
        arrNegatif.append(kata)
    
    print('arrPositif: ', arrPositif)
    print('arrNegatif: ', arrNegatif)
    
    return arrPositif, arrNegatif

# =================================================================================================================
# ============================================== SENTIMENT ANALYSIS ===============================================
# =================================================================================================================

# -------------COUNT SENTIMENT SCORE-------------
def sentimentScore(hasilPositif, hasilNegatif):
    countPositif = 0
    countNegatif = 0
    
    for arr in hasilPositif:
        countPositif = countPositif + arr[1]
    for arr in hasilNegatif:
        countNegatif = countNegatif + arr[1]

    
    # print('countPositif: ', countPositif)
    # print('countNegatif: ', countNegatif)

    sentimentScore = countNegatif + countPositif
 
    return countPositif, countNegatif, sentimentScore

# -------------CHECK SENTIMENT RESULT-------------
def cekSentimen(nilaiSentimen):
    hasil = ''
    if (nilaiSentimen > 0):
        hasil = 'POSITIVE'
    elif (nilaiSentimen < 0):
        hasil = 'NEGATIVE'
    elif (nilaiSentimen == 0):
        hasil = 'NEUTRAL'
    
    return hasil

# =================================================================================================================
# ================================================= MAIN PROGRAM ==================================================
# =================================================================================================================
def counting_sentiment(research_project):
    # dataset = workbook
    answers = models.ResearchAnswer.objects.filter(research_project=research_project)
    total_hasil_sentiment = 0
    total_hasil_positif = 0
    total_hasil_negatif = 0
    for answer in answers:
        dataStatis = answer.answer
        hasilToken = tokenization(dataStatis)
        ngramPositif, ngramNegatif, ngram1 = sentiWordBeforeStem(hasilToken)
        hasilStem = stemmingWord(ngram1)
        hasilNoPuct = punctuationRemoval(hasilStem)
        hasilStopWord = stopwordRemoval(hasilNoPuct)
        hasilPraprosesCoding = hasilStopWord

        # print('hasil praposes: ', hasilPraprosesCoding)
        
        hasilPositif, hasilNegatif = sentiWordAfterStem(hasilPraprosesCoding, ngramPositif, ngramNegatif)
        countPositif, countNegatif, hasilSentimen = sentimentScore(hasilPositif, hasilNegatif)

        total_hasil_positif = total_hasil_positif + countPositif
        total_hasil_negatif = total_hasil_negatif + countNegatif
        total_hasil_sentiment = total_hasil_sentiment + hasilSentimen

    # print('hasil total positif: ', total_hasil_positif)
    # print('hasil total negatif: ', total_hasil_negatif)
    # print('hasil total sentimen score: ', total_hasil_sentiment)
    nilai = cekSentimen(total_hasil_sentiment)
    return total_hasil_positif, total_hasil_negatif, total_hasil_sentiment, nilai



# def converttoexcel(pk):
#      research_project = models.ResearchProject.objects.get(pk=pk)
#      answers_query = models.ResearchAnswer.objects.filter(research_project=pk)
#     #  response = HttpResponse(
#     #       content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     #  )
#     #  response['Content-Disposition'] = 'attachment; filename=answer.xlsx'
#      workbook = Workbook()
#      # Get active worksheet/tab
#      worksheet = workbook.active
#      worksheet.title = 'answers'
#      # Define the titles for columns
#      columns = [
#           'answers',
#      ]
#      row_num = 0
#      # Assign the titles for each cell of the header
#      # for col_num, column_title in enumerate(columns, 1):
#      #     cell = worksheet.cell(row=row_num, column=col_num)
#      #     cell.value = column_title
#      # Iterate through all movies
#      for researchanswer in answers_query:
#           row_num += 1
#      # Define the data for each cell in the row 
#           row = [
#                researchanswer.answer,
#           ]
#      # Assign the data for each cell of the row 
#           for col_num, cell_value in enumerate(row, 1):
#                cell = worksheet.cell(row=row_num, column=col_num)
#                cell.value = cell_value
#      sheet=workbook.active
#      max_row = sheet.max_row
#      max_column = sheet.max_column
#      for row_num in range(1, max_row+1):
#           for col_num in range(1, max_column+1):
#                cell=sheet.cell(row=row_num, column=col_num)
#                print(cell.value,end=' | ')
#           print('\n')
#      return workbook
     
